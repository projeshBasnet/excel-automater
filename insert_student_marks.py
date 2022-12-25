from pathlib import Path
from helper_function import find_excel_sheet, find_pdf_file

from openpyxl import load_workbook
from pathlib import Path
import numpy as np
import random
import json

# path = Path(folder_location)
# cell_with_marks_to_be_changed = {"D11":1,"D12":1,"D12":1,"D14":1,"D15":2,"D17":2,"D18":2,"D22":2}



def add_formula_to_cells(excel_file_path):
    print("Adding formula to the excel sheet")
    workbook = load_workbook(excel_file_path)
    sheet1 = workbook["Grading Sheet"]
    # adding sum formula to cell D11
    sheet1["D10"] = "=SUM(D11:D15)"
    sheet1["D16"] = "=SUM(D17:D18)"
    # adding grand total formula in toal marks
    sheet1["D24"] = "=SUM(D9,D10,D16,D19,D20,D21,D22)"

    # now adding formula to Result Sheet
    first_sheet_name = "Grading Sheet"
    sheet2 = workbook["Result"]

    sheet2["C7"] = f"='{first_sheet_name}'!D9"
    sheet2["C9"] = f"='{first_sheet_name}'!D11"
    sheet2["C10"] = f"='{first_sheet_name}'!D12"
    sheet2["C11"] = f"='{first_sheet_name}'!D13"
    sheet2["C12"] = f"='{first_sheet_name}'!D14"
    sheet2["C13"] = f"='{first_sheet_name}'!D15"
    sheet2["C15"] = f"='{first_sheet_name}'!D17"
    sheet2["C16"] = f"='{first_sheet_name}'!D18"
    sheet2["C18"] = f"='{first_sheet_name}'!D19"
    sheet2["C20"] = f"='{first_sheet_name}'!D20"
    sheet2["C22"] = f"='{first_sheet_name}'!D21"
    sheet2["C24"] = f"='{first_sheet_name}'!D22"
    sheet2["C26"] = f"='{first_sheet_name}'!D24"
    sheet2["E26"] = f"=C26"

    # sheet2["C13"] = "=SUM(C6,C7,C8,C9,C10,C11)"
    # sheet2["E13"] = "=C13"


    workbook.save(excel_file_path)
    # workbook.close()
    print(f"Sucessfully written the formula to sheeet")





# def change_excel_marks(excel_file_path):
#     workbook = load_workbook(excel_file_path, data_only=True)
#     sheet = workbook["Grading Sheet"]
#     for key, value in cell_with_marks_to_be_changed.items():
#         print(f"Changing the marks of cell: {key}")
#         sheet[key] = sheet[key].value - value
#     workbook.save(excel_file_path)
#     workbook.close()


# main code to iter through directories and writing into excel
# for p in path.iterdir():
#     for file in p.rglob("*.xlsx"):
#         print(f"ExcelFile is : {file}")
        
       
#         print("ABout to Change marks  in the cells")
#         change_excel_marks(file)

#         print("ABout to write formula to cells")
#         add_formula_to_cells(file)
#         print(f"sucessfully changed marks of sheet: {file}")

       

print(f"completed changing marks of student in that folder")


def add_marks_on_student_sheet(excel_file, marks_info):
    print("inside add_marks_on_student_sheet")
    workbook = load_workbook(excel_file)
    sheet = workbook["Grading Sheet"]
    print(f"marks_info: {marks_info}")
    for key, value in marks_info.items():
        print(f"value is: {value}")
        value = json.loads(value)
        print(f"Changing the marks of cell: {key}")
        """
        it creates a marks value from a given lists randomly within a given range having spacing of 0.5
        """

        marks_value = random.choice(np.arange(value[0],value[1]+1,0.5)) 
        print(f"marks_value: {marks_value}")

        sheet[key] = marks_value

    workbook.save(excel_file)
    workbook.close()

    

def handle_insert_student_marks(folder_name,marks_info):
    for path in Path(folder_name).iterdir():
        excel_file = find_excel_sheet(path)
        add_marks_on_student_sheet(excel_file,marks_info)


        
