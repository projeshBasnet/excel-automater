import openpyxl
import os
from pathlib import Path


start_depth = 11
end_depth = 49
code_column_label = "B"



letters = ['H','I','J','K','L','M','N','O','P','Q']

def find_row_number(student_code, ws):
    print("inside find row")
    print(type(student_code))
    print("student code", student_code)
    row_no = 0
    for i in range(start_depth, end_depth+1):
        print("inside for loop")
        print("equal val",ws[f"{code_column_label}{i}"].value)
        if student_code == ws[f"{code_column_label}{i}"].value:
            row_no = i
            break
        
    return row_no




# write to google docs
def write_to_google_docs(student_id, marks_array):
    new_excel_path = parent_path / "excel_files" 
    file_path = f"{new_excel_path}/CS4051NT 2021-22 SEM2 Result.xlsx"


    print("path",os.path.dirname(os.path.realpath(__file__)))
    wb = openpyxl.load_workbook(f"{file_path}")
    ws = wb["001"]
    row_num = int(find_row_number(student_id, ws))

    for idx, letter in enumerate(letters):
        print(f"marks_array[idx]: {type(marks_array[idx])}")
        print(f'ws[f"{letter}{row_num}"]: {type(ws[f"{letter}{row_num}"].value)})')
        ws[f"{letter}{row_num}"].value = marks_array[idx]

    wb.save(file_path)

    print("..............")
    print(f"completed writing into first row {row_num}")

    # get_values_of_row(student_id)

student_column = "D"
student_marks_arr = [f"{student_column}9",f"{student_column}10",f"{student_column}11", f"{student_column}12",f"{student_column}13",f"{student_column}14",f"{student_column}20",f"{student_column}21",f"{student_column}22",f"{student_column}23"]

parent_path = Path(__file__).parents[1]
excel_path = parent_path / "excel_files" / "sample_excel"
path = Path(excel_path)

for p in path.iterdir():
    student_id = int(os.path.basename(p).split(" ")[0])

    for file in p.rglob("*.xlsx"):
        print("excel_file", file)
        wb = openpyxl.load_workbook(file, data_only=True)
        print("workbook", wb)
        ws = wb["Grading Sheet"]
        marks_array = []
        for idx in student_marks_arr:
            marks_array.append(int(ws[idx].value))
        

    write_to_google_docs(student_id, marks_array)

print("completed writing into sheet now!!!!!!")








   
        



