from asyncore import write
import openpyxl
from pathlib import Path
import os
import logging


from win32com import client
from execute_pdf import merge_excel_sheet_to_pdf
win32_excel = client.Dispatch("Excel.Application")




class WriteToExcel:
    def __init__(self, file_name, sheet_name, start_depth, end_depth, column_start_range, column_end_range, student_id_column = "B") -> None:
        self.file_name = file_name
        self.sheet_name = sheet_name
        self.start_depth = start_depth
        self.end_depth = end_depth
        self.column_start_range = column_start_range
        self.column_end_range = column_end_range
        self.iterating_range = [i for i in range(self.start_depth, self.end_depth+1)]
        print(f"self.iterating_range: {self.iterating_range}")
        self.student_id_column = student_id_column
        self.column_array_range = self.create_column_array()
        self.work_book = openpyxl.load_workbook(self.file_name)
        self.sheet = self.work_book[self.sheet_name]
        self.sheet_marks_position = self.create_sheet_marks_position()

        print(f"sheet_marks_position: {self.sheet_marks_position}")

    
    def range_char(self,start, stop):
        return (chr(n) for n in range(ord(start), ord(stop) + 1))

    def create_column_array(self):
        my_array = []
        for char in self.range_char(self.column_start_range, self.column_end_range):
            my_array.append(char)
        return my_array
    
    def create_sheet_marks_position(self):
        my_dict = {}
        
        for letter in self.column_array_range:
            value = self.sheet[f"{letter}{self.start_depth-1}"].value.split("(")[0].strip()
            my_dict[value] = f"{letter}"

        return my_dict
    
    def find_student_row_number(self, student_id):
        print("inside find_student_row_number")
        row_no = 0
        print(f"student_id: {student_id}")
        for i in self.iterating_range:
            print("Inside for loop")
            if student_id == self.sheet[f"{self.student_id_column}{i}"].value:
                row_no = i
                self.iterating_range.remove(i)
                print(f"changed iterating range: {self.iterating_range}")

                break
          
        print(f"student row number about to be returned: {row_no}")
        return row_no



class ReadFromExcel:
    def __init__(self, folder_path, marks_dict, sheet_name, excel_object,merge_excel) -> None:
        self.folder_path = folder_path 
        self.marks_dict = marks_dict
        self.sheet_name = sheet_name
        self.excel_object = excel_object
        self.merge_excel = merge_excel

        # create a logger
        desktop = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
        logging.basicConfig(filename = f'{desktop}/log_message.txt',encoding ='utf-8', level=logging.DEBUG, filemode='w',format='%(levelname)s:%(message)s')
        self.logger = logging.getLogger("excel_writer")
        self.logger.setLevel(logging.DEBUG)

        ch = logging.StreamHandler()
        ch.setLevel(logging.DEBUG)
        formatter = logging.Formatter('%(levelname)s - %(message)s')
        ch.setFormatter(formatter)
        self.logger.addHandler(ch)
        



        print(f"self.logger: {self.logger}")
       
    def read_folder(self):
        path = Path(self.folder_path)
        print(f"path: {path}")
        """
        Reading the each folder of student and taking the student id
        """
        for p in path.iterdir():
            try:
                student_id = int(os.path.basename(p).split()[0])
            except:
                self.logger.error(f"Folder of student {os.path.basename(p).split()[0]} does not contain student id")
                continue
            self.read_from_excel_file(student_id, p)

       
            
    def read_from_excel_file(self, student_id,p):
        print("inside read_from_excel_file")
        excel_file = None
        import re
        print(f"main folder is:",str(p))
        for file in p.rglob("*.xlsx"):
            print(f"excel file: {file}")
            if not re.match(r"[^~$].*",str(file)):
                continue
            else:
                print(f"file in else: {file}")
                excel_file = file
                wb = openpyxl.load_workbook(file, data_only=True,read_only=True)
                ws = wb[self.sheet_name]
                marks_value_dict = {}
                for key, value in self.marks_dict.items():
                    # has used this try block to convert 0 string into integer
                    try:
                        marks_value_dict[key] = float(ws[value].value)
                    except ValueError:
                        marks_value_dict[key] = 0

                self.write_to_main_docs(student_id, marks_value_dict)
            
            
                print(f"active workbook: {wb.active}")
                wb.close()
                break
            
         

        if self.merge_excel:
            for file in p.rglob("*.pdf"):
                merge_excel_sheet_to_pdf(str(p),excel_file=excel_file,pdf_file=file)
                break
                
    def write_to_main_docs(self,student_id, marks_value_dict):
        print(f"writing to main docs of id {student_id}")
        print(f"marks_value_dict: {marks_value_dict}")
        print(f"sheet marks position: {self.excel_object.sheet_marks_position}")
 
        if len(marks_value_dict.keys()) == len(self.excel_object.sheet_marks_position.keys()):
            # used try block to check if the row number  of student is correct or not
            
            student_row_num = self.excel_object.find_student_row_number(student_id)
            if student_row_num >0:
                for key, value in marks_value_dict.items():
                    print(f"student_row_num: {student_row_num}")
                    if key in self.excel_object.sheet_marks_position:
                        self.excel_object.sheet[f"{self.excel_object.sheet_marks_position[key]}{student_row_num}"] = value
                self.excel_object.work_book.save(self.excel_object.file_name)
                print(f"sucessfully saved file")
            else:
                self.logger.error(f"Error in london met id of student {student_id}")
        
        else:
            print(f"Marks dictonary is not so valid with")
                
                


